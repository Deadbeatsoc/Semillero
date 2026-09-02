# -*- coding: utf-8 -*-
"""
Geocodifica los siniestros reales (Excel Secretaria de Movilidad) usando
interseccion de vias nombradas de OpenStreetMap.

Estrategia (sin rate-limit de geocoder):
  1. Descarga UNA sola vez todas las vias nombradas de Villavicencio (Overpass).
  2. Indexa por nombre canonico ("calle 15", "carrera 22", ...).
  3. Para cada direccion "CALLE N CON CARRERA M" calcula la interseccion
     geometrica de las dos vias. Fallback: vertice mas cercano entre ambas.
  4. Escribe backend/data/real_accidents.csv (solo filas geocodificadas).

Uso:  python scripts/geocode_real_accidents.py
"""
import os, re, csv, json, math, unicodedata, urllib.request, urllib.parse
from datetime import datetime, time as dtime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "..", "data")
EXCEL_PATH = os.environ.get(
    "REAL_EXCEL_PATH",
    r"C:\Users\Omarl\Documents\Universidad\Semillero\01. Siniestralidad_2023-2023(Informe)LIMPIAR.xlsx",
)
SHEETS = ["GENERAL ACCIDENTALIDAD 2022", "GENERAL ACCIDENTALIDAD 2023"]
BBOX = (4.01, -73.78, 4.27, -73.48)  # south, west, north, east
WAYS_CACHE = os.path.join(DATA_DIR, "osm_named_ways.json")
OUT_CSV = os.path.join(DATA_DIR, "real_accidents.csv")
OVERPASS_URL = os.environ.get("OVERPASS_URL", "https://overpass-api.de/api/interpreter")

# ---- columnas (0-based, con columna inicial en blanco) ----
C_FECHA, C_HORA, C_SEV, C_TOTHER, C_DIR, C_VEH = 1, 5, 9, 13, 14, 15


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_text(s):
    return re.sub(r"\s+", " ", strip_accents(str(s or "")).strip().lower())


# Prefijos de via -> tipo canonico
TIPO_MAP = {
    "calle": "calle", "cll": "calle", "cl": "calle", "cale": "calle",
    "carrera": "carrera", "cra": "carrera", "cr": "carrera", "kra": "carrera", "kr": "carrera", "k": "carrera",
    "avenida": "avenida", "av": "avenida", "ave": "avenida",
    "diagonal": "diagonal", "diag": "diagonal", "dg": "diagonal",
    "transversal": "transversal", "trans": "transversal", "tv": "transversal", "tr": "transversal",
    "circunvalar": "circunvalar",
}


def canon_via(tipo, numero):
    """Normaliza ('CRA','22A') -> 'carrera 22a'."""
    t = TIPO_MAP.get(tipo, tipo)
    n = numero.lower().replace(" ", "")
    return f"{t} {n}"


def canon_osm_name(name):
    """Convierte el name de OSM a clave canonica (y variantes utiles)."""
    n = norm_text(name)
    n = n.replace("avenida calle", "calle").replace("avenida carrera", "carrera")
    m = re.match(r"^(calle|carrera|avenida|diagonal|transversal|circunvalar)\s+([0-9]+[a-z]?)", n)
    if not m:
        return None
    return f"{m.group(1)} {m.group(2)}"


VIA_TOKEN = re.compile(
    r"\b(calle|cll|cl|cale|carrera|cra|cr|kra|kr|k|avenida|av|ave|diagonal|diag|dg|transversal|trans|tv|tr|circunvalar)\s*#?\s*([0-9]{1,3}[a-z]?)",
    re.I,
)


def parse_two_vias(direccion):
    """Extrae hasta dos vias canonicas de una direccion."""
    txt = norm_text(direccion)
    vias = []
    for m in VIA_TOKEN.finditer(txt):
        tipo = TIPO_MAP.get(m.group(1).lower(), m.group(1).lower())
        via = canon_via(m.group(1).lower(), m.group(2))
        if via not in vias:
            vias.append(via)
        if len(vias) >= 2:
            break
    return vias


def fetch_named_ways():
    if os.path.exists(WAYS_CACHE):
        with open(WAYS_CACHE, "r", encoding="utf-8") as fh:
            return json.load(fh)
    s, w, n, e = BBOX
    q = f'[out:json][timeout:120];way["highway"]["name"]({s},{w},{n},{e});out geom;'
    data = urllib.parse.urlencode({"data": q}).encode()
    req = urllib.request.Request(OVERPASS_URL, data=data, headers={"User-Agent": "semillero-villavicencio/1.0"})
    print("Descargando vias nombradas de OSM (una sola vez)...")
    with urllib.request.urlopen(req, timeout=180) as r:
        payload = json.load(r)
    ways = []
    for el in payload.get("elements", []):
        if el.get("type") != "way":
            continue
        geom = el.get("geometry") or []
        name = (el.get("tags") or {}).get("name")
        if name and len(geom) >= 2:
            ways.append({"name": name, "geom": [(p["lat"], p["lon"]) for p in geom]})
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(WAYS_CACHE, "w", encoding="utf-8") as fh:
        json.dump(ways, fh)
    print(f"  {len(ways)} vias nombradas cacheadas en {WAYS_CACHE}")
    return ways


def build_index(ways):
    idx = {}
    for way in ways:
        key = canon_osm_name(way["name"])
        if not key:
            continue
        idx.setdefault(key, []).append(way["geom"])
    return idx


def seg_intersection(p1, p2, p3, p4):
    """Interseccion de segmentos p1p2 y p3p4 (lat,lon como plano local)."""
    x1, y1 = p1[1], p1[0]
    x2, y2 = p2[1], p2[0]
    x3, y3 = p3[1], p3[0]
    x4, y4 = p4[1], p4[0]
    den = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)
    if abs(den) < 1e-12:
        return None
    t = ((x1 - x3) * (y3 - y4) - (y1 - y3) * (x3 - x4)) / den
    u = ((x1 - x3) * (y1 - y2) - (y1 - y3) * (x1 - x2)) / den
    if 0 <= t <= 1 and 0 <= u <= 1:
        return (y1 + t * (y2 - y1), x1 + t * (x2 - x1))  # (lat, lon)
    return None


def intersect_ways(linesA, linesB):
    for la in linesA:
        for i in range(len(la) - 1):
            for lb in linesB:
                for j in range(len(lb) - 1):
                    pt = seg_intersection(la[i], la[i + 1], lb[j], lb[j + 1])
                    if pt:
                        return pt
    return None


def nearest_vertices_mid(linesA, linesB):
    best = None
    bestd = float("inf")
    for la in linesA:
        for a in la:
            for lb in linesB:
                for b in lb:
                    d = (a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2
                    if d < bestd:
                        bestd = d
                        best = ((a[0] + b[0]) / 2.0, (a[1] + b[1]) / 2.0)
    # Solo aceptamos si las vias estan razonablemente cerca (~250 m)
    if best and math.sqrt(bestd) < 0.0025:
        return best
    return None


def in_bbox(lat, lon):
    s, w, n, e = BBOX
    return s <= lat <= n and w <= lon <= e


def map_severity(sev_raw, total_heridos):
    s = norm_text(sev_raw)
    if s.startswith("solo"):  # SOLO DANOS
        return "baja"
    if s.startswith("herido"):
        try:
            th = int(float(total_heridos))
        except (TypeError, ValueError):
            th = 1
        return "alta" if th >= 3 else "media"
    return "media"


def parse_datetime(fecha, hora):
    d = None
    if isinstance(fecha, datetime):
        d = fecha
    else:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                d = datetime.strptime(str(fecha).strip(), fmt)
                break
            except (ValueError, TypeError):
                continue
    if d is None:
        return None
    hh, mm, ss = 12, 0, 0
    if isinstance(hora, dtime):
        hh, mm, ss = hora.hour, hora.minute, hora.second
    elif hora:
        m = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", str(hora).strip())
        if m:
            hh, mm = int(m.group(1)), int(m.group(2))
            ss = int(m.group(3) or 0)
    return d.replace(hour=hh, minute=mm, second=ss)


def main():
    ways = fetch_named_ways()
    index = build_index(ways)
    print(f"Indice de vias canonicas: {len(index)} nombres unicos")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rows_out = []
    stats = {"total": 0, "geocoded_intersection": 0, "geocoded_nearest": 0, "no_streets": 0, "no_match": 0, "no_date": 0}

    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  (hoja ausente: {sheet})")
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or len(r) <= C_VEH:
                continue
            direccion = r[C_DIR]
            fecha = r[C_FECHA]
            if not direccion or not fecha:
                continue
            stats["total"] += 1

            dt = parse_datetime(fecha, r[C_HORA])
            if dt is None:
                stats["no_date"] += 1
                continue

            vias = parse_two_vias(direccion)
            if len(vias) < 2:
                stats["no_streets"] += 1
                continue
            linesA = index.get(vias[0])
            linesB = index.get(vias[1])
            if not linesA or not linesB:
                stats["no_match"] += 1
                continue

            pt = intersect_ways(linesA, linesB)
            method = "geocoded_intersection"
            if pt is None:
                pt = nearest_vertices_mid(linesA, linesB)
                method = "geocoded_nearest"
            if pt is None or not in_bbox(pt[0], pt[1]):
                stats["no_match"] += 1
                continue
            stats[method] += 1

            hour = dt.hour
            period = "dia" if 6 <= hour < 18 else "noche"
            severity = map_severity(r[C_SEV], r[C_TOTHER])
            veh = norm_text(r[C_VEH]).upper() if r[C_VEH] else ""
            desc = f"Siniestro real Secretaria Movilidad ({veh})" if veh else "Siniestro real Secretaria Movilidad"
            rows_out.append({
                "occurred_at": dt.strftime("%Y-%m-%d %H:%M:%S"),
                "severity": severity,
                "weather": "desconocido",
                "period": period,
                "latitude": round(pt[0], 7),
                "longitude": round(pt[1], 7),
                "description": desc[:240],
                "address": str(direccion)[:240],
            })
    wb.close()

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["occurred_at", "severity", "weather", "period", "latitude", "longitude", "description", "address"])
        writer.writeheader()
        writer.writerows(rows_out)
    # Salida JSON (consumida por el importador de Node, robusta ante comas/comillas).
    with open(os.path.join(DATA_DIR, "real_accidents.json"), "w", encoding="utf-8") as fh:
        json.dump(rows_out, fh, ensure_ascii=False)

    geocoded = stats["geocoded_intersection"] + stats["geocoded_nearest"]
    print("\n=== Resultado geocodificacion ===")
    print(f"  Total filas con direccion+fecha : {stats['total']}")
    print(f"  Geocodificadas (interseccion)   : {stats['geocoded_intersection']}")
    print(f"  Geocodificadas (vertice cercano): {stats['geocoded_nearest']}")
    print(f"  Total geocodificadas            : {geocoded}  ({(geocoded/max(1,stats['total'])*100):.1f}%)")
    print(f"  Sin 2 vias parseables           : {stats['no_streets']}")
    print(f"  Vias no encontradas en OSM      : {stats['no_match']}")
    print(f"  Sin fecha valida                : {stats['no_date']}")
    print(f"  CSV escrito en: {OUT_CSV}")


if __name__ == "__main__":
    main()
